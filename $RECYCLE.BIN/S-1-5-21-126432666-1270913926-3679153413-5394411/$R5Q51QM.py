import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def get_relative_files(base_path):
    """Get all files in a directory recursively, returning relative paths."""
    files = set()
    base = Path(base_path)
    if not base.exists():
        return files
    
    for file in base.rglob('*'):
        if file.is_file():
            relative_path = file.relative_to(base)
            files.add(str(relative_path).replace('\\', '/'))
    return files

def main():
    workspace = Path(r'd:\Repo\cookbook\cookbook')
    
    # Get files from main docs folder
    main_docs = workspace / 'docs'
    main_files = get_relative_files(main_docs)
    
    # Get versioned docs folders
    versioned_docs_dir = workspace / 'versioned_docs'
    if not versioned_docs_dir.exists():
        print("No versioned_docs folder found!")
        return
    
    # Get files from version-2506 and version-2512
    version_2506_dir = versioned_docs_dir / 'version-2506'
    version_2512_dir = versioned_docs_dir / 'version-2512'
    
    if not version_2506_dir.exists():
        print("version-2506 not found!")
        return
    if not version_2512_dir.exists():
        print("version-2512 not found!")
        return
    
    version_2506_files = get_relative_files(version_2506_dir)
    version_2512_files = get_relative_files(version_2512_dir)
    
    # Exclude PNG files from the comparison
    def exclude_png_files(files):
        return {file for file in files if not file.endswith('.png')}

    version_2506_files = exclude_png_files(version_2506_files)
    version_2512_files = exclude_png_files(version_2512_files)
    
    print(f"Found {len(main_files)} files in main docs")
    print(f"Found {len(version_2506_files)} files in version-2506")
    print(f"Found {len(version_2512_files)} files in version-2512\n")
    
    # Exclude main docs from the comparison
    all_files = version_2506_files | version_2512_files
    
    # Filter folders to include only the specified ones
    def filter_folders(files, allowed_folders):
        return {file for file in files if any(file.startswith(folder) for folder in allowed_folders)}

    allowed_folders = [
        "Documentation/",
        "Product Integration Documentation/",
        "CTCX Intro/"
    ]

    version_2506_files = filter_folders(version_2506_files, allowed_folders)
    version_2512_files = filter_folders(version_2512_files, allowed_folders)

    print(f"Filtered {len(version_2506_files)} files in version-2506")
    print(f"Filtered {len(version_2512_files)} files in version-2512\n")
    
    # Filter out files with identical content
    def filter_identical_files(files_2506, files_2512, base_path_2506, base_path_2512):
        unique_files_2506 = set()
        unique_files_2512 = set()

        for file in files_2506 | files_2512:
            file_2506 = base_path_2506 / file
            file_2512 = base_path_2512 / file

            if file in files_2506 and file in files_2512:
                if file_2506.read_text(errors='ignore') != file_2512.read_text(errors='ignore'):
                    unique_files_2506.add(file)
                    unique_files_2512.add(file)
            elif file in files_2506:
                unique_files_2506.add(file)
            elif file in files_2512:
                unique_files_2512.add(file)

        return unique_files_2506, unique_files_2512

    version_2506_files, version_2512_files = filter_identical_files(
        version_2506_files, version_2512_files, version_2506_dir, version_2512_dir
    )

    print(f"Filtered {len(version_2506_files)} files in version-2506 (unique)")
    print(f"Filtered {len(version_2512_files)} files in version-2512 (unique)\n")
    
    # Get all unique files across all versions
    all_files = version_2506_files | version_2512_files
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Docs Comparison"
    
    # Set up headers
    headers = ["File Path", "In Version-2506", "In Version-2512", "Status"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Define status messages
    def get_status(in_2506, in_2512):
        if in_2506 and in_2512:
            return "In both versions"
        elif in_2506:
            return "Only in v2506"
        elif in_2512:
            return "Only in v2512"
        else:
            return "Unknown"
    
    # Define fills for different statuses
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Populate data
    for file_path in sorted(all_files):
        in_2506 = file_path in version_2506_files
        in_2512 = file_path in version_2512_files
        status = get_status(in_2506, in_2512)
        
        row = [
            file_path,
            "Yes" if in_2506 else "No",
            "Yes" if in_2512 else "No",
            status
        ]
        ws.append(row)
        
        # Apply color coding to the last row
        current_row = ws.max_row
        if not in_2506 and not in_2512:
            # Files not in any version - highlight in yellow/orange
            ws[f"A{current_row}"].fill = yellow_fill
            ws[f"E{current_row}"].fill = yellow_fill
        elif in_2506 and in_2512:
            # Files in both versions - light green
            ws[f"E{current_row}"].fill = green_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add filters
    ws.auto_filter.ref = ws.dimensions
    
    # Save the file
    output_file = workspace / "docs_comparison.xlsx"
    wb.save(output_file)
    
    print(f"Excel file created: {output_file}")
    print("\nSummary:")
    print(f"  - Total unique files: {len(all_files)}")
    print(f"  - Files in main docs: {len(main_files)}")
    print(f"  - Files in version-2506: {len(version_2506_files)}")
    print(f"  - Files in version-2512: {len(version_2512_files)}")
    print(f"  - Files in v2506 but not in main: {len(version_2506_files - main_files)}")
    print(f"  - Files in v2512 but not in main: {len(version_2512_files - main_files)}")
    print(f"  - Files in both v2506 and v2512 but not in main: {len((version_2506_files & version_2512_files) - main_files)}")

if __name__ == '__main__':
    main()
