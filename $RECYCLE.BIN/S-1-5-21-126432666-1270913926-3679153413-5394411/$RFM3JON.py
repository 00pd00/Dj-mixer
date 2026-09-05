import os
import re
import urllib.parse
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────
BASE_URL = "https://ctcx.code.siemens.io/cookbook/docs/2606"
DOCS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Cookbook_Pages.xlsx")

# Columns A-D are auto-generated; columns E onward are user-managed
AUTO_COLS = 4  # Category, Sub-Section, Link, Page Name

# Extra header columns (E–L) to create if file doesn't exist yet
EXTRA_HEADERS = [
    "Description",
    "Redo Post Deploy after TC \nUpdate Existing Env with\n New Patch",
    "Expected time to complete",
    " Oracle-Ent(BYOL)-HA\nPremium-tc2512.2025110700\n25121711-prd\nECA: 500126810",
    " PostSQL-SA\nPremium-tc2512.2025110700\n25121811-prd\nECA: 500126811",
    "Premium-tc2512.2025110700 25121711-prd4 \nECA:500126812",
    "PRs reported",
    "Notes",
]


def strip_numeric_prefix(name):
    """Remove leading numeric prefix like '000_', '010_' etc."""
    return re.sub(r"^\d+[_\s]+", "", name)


# ── Read existing user data (columns E+) keyed by Link URL ────
existing_data = {}  # url -> {col_letter: value, ...}
extra_headers_from_file = []
if os.path.exists(OUTPUT_FILE):
    try:
        old_wb = load_workbook(OUTPUT_FILE)
        old_ws = old_wb.active
        # Read extra headers from existing file
        for col_idx in range(AUTO_COLS + 1, old_ws.max_column + 1):
            val = old_ws.cell(row=1, column=col_idx).value
            if val is not None:
                extra_headers_from_file.append(val)
        # Read user data keyed by the Link URL (column C hyperlink)
        for row in old_ws.iter_rows(min_row=2, max_row=old_ws.max_row):
            link_cell = row[2]  # column C = Link
            url_key = link_cell.hyperlink.target if link_cell.hyperlink else None
            if not url_key:
                continue
            row_data = {}
            for col_idx in range(AUTO_COLS + 1, old_ws.max_column + 1):
                cell = row[col_idx - 1]
                if cell.value is not None:
                    row_data[col_idx] = cell.value
            if row_data:
                existing_data[url_key] = row_data
        old_wb.close()
        print(f"Read {len(existing_data)} rows of user data from existing file")
    except Exception as e:
        print(f"Warning: could not read existing file: {e}")

# Use headers from existing file if available, otherwise defaults
if not extra_headers_from_file:
    extra_headers_from_file = EXTRA_HEADERS

# ── Collect all .md pages grouped by section ──────────────────
sections = OrderedDict()
for root, dirs, files in os.walk(DOCS_DIR):
    dirs.sort()
    files.sort()
    for f in sorted(files):
        if not f.endswith(".md"):
            continue
        full = os.path.join(root, f)
        rel = os.path.relpath(full, DOCS_DIR)
        parts = rel.replace("\\", "/").split("/")

        # Top-level section (category)
        section = strip_numeric_prefix(parts[0]) if len(parts) > 1 else "Root"

        # Sub-section: everything between section and filename
        sub_parts = parts[1:-1]
        sub_section = " / ".join(strip_numeric_prefix(s) for s in sub_parts) if sub_parts else ""

        # Page name (without .md, without numeric prefix)
        page_name = strip_numeric_prefix(os.path.splitext(parts[-1])[0])

        # Build URL: encode each segment (spaces -> %20), strip .md from last segment
        url_segments = []
        for i, p in enumerate(parts):
            seg = os.path.splitext(p)[0] if i == len(parts) - 1 else p
            url_segments.append(urllib.parse.quote(seg, safe="_-"))
        url = BASE_URL + "/" + "/".join(url_segments)

        sections.setdefault(section, []).append((sub_section, page_name, url))

# ── Create Excel workbook ─────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Cookbook Pages"

# Styling
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(bold=True, size=11)
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
link_font = Font(color="0563C1", underline="single")

# Column order: A=Category, B=Sub-Section, C=Link, D=Page Name, E+=user columns
all_headers = ["Category", "Sub-Section", "Link", "Page Name"] + extra_headers_from_file
ws.append(all_headers)
for col_idx in range(1, len(all_headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

subsection_font = Font(bold=False, size=10)
subsection_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")

# Data rows — grouped by section, then by sub-section (both merged)
for section, pages in sections.items():
    section_start = ws.max_row + 1

    # Group pages by sub-section (preserving order)
    sub_groups = OrderedDict()
    for sub_section, page_name, url in pages:
        sub_groups.setdefault(sub_section, []).append((page_name, url))

    for sub_section, sub_pages in sub_groups.items():
        sub_start = ws.max_row + 1

        for page_name, url in sub_pages:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1).border = thin_border                       # A: Category (merged later)
            ws.cell(row=row_num, column=2).border = thin_border                       # B: Sub-Section (merged later)
            link_cell = ws.cell(row=row_num, column=3, value="Link")                  # C: Link
            link_cell.hyperlink = url
            link_cell.font = link_font
            link_cell.border = thin_border
            ws.cell(row=row_num, column=4, value=page_name).border = thin_border      # D: Page Name

            # Restore user data from existing file
            if url in existing_data:
                for col_idx, val in existing_data[url].items():
                    ws.cell(row=row_num, column=col_idx, value=val)

            # Add borders for extra columns
            for col_idx in range(AUTO_COLS + 1, len(all_headers) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border

        sub_end = ws.max_row

        # Merge sub-section column for all rows of this sub-section
        if sub_end > sub_start:
            ws.merge_cells(start_row=sub_start, start_column=2, end_row=sub_end, end_column=2)

        sub_cell = ws.cell(row=sub_start, column=2, value=sub_section)
        sub_cell.font = subsection_font
        sub_cell.fill = subsection_fill
        sub_cell.alignment = Alignment(vertical="center", wrap_text=True)
        sub_cell.border = thin_border

    section_end = ws.max_row

    # Merge category column for all rows of this section
    if section_end > section_start:
        ws.merge_cells(start_row=section_start, start_column=1, end_row=section_end, end_column=1)

    section_cell = ws.cell(row=section_start, column=1, value=section)
    section_cell.font = section_font
    section_cell.fill = section_fill
    section_cell.alignment = Alignment(vertical="center", wrap_text=True)
    section_cell.border = thin_border

# Auto-fit column widths (capped at 60)
for col in ws.columns:
    max_len = max(len(str(c.value or "").split("\n")[0]) for c in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

# Freeze header row and add auto-filter
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)
total = sum(len(pages) for pages in sections.values())
print(f"Created {OUTPUT_FILE} with {total} pages across {len(sections)} sections")
